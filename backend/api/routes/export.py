"""
Export endpoints: JSON, CSV, Excel.
GET /api/audits/{audit_id}/export/json
GET /api/audits/{audit_id}/export/csv
GET /api/audits/{audit_id}/export/xlsx
"""

import csv
import io
import json
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.db.database import get_session
from backend.db.models import AuditRecord, FindingRecord

router = APIRouter(prefix="/api/audits", tags=["export"])


async def _get_audit_and_findings(audit_id: str, session: AsyncSession):
    result = await session.execute(select(AuditRecord).where(AuditRecord.id == audit_id))
    audit = result.scalar_one_or_none()
    if not audit:
        raise HTTPException(status_code=404, detail="Audit not found")

    f_result = await session.execute(
        select(FindingRecord).where(FindingRecord.audit_id == audit_id)
    )
    findings = f_result.scalars().all()
    return audit, findings


@router.get("/{audit_id}/export/json")
async def export_json(audit_id: str, session: AsyncSession = Depends(get_session)):
    audit, findings = await _get_audit_and_findings(audit_id, session)

    data = {
        "audit": {
            "id": audit.id,
            "company_name": audit.company_name,
            "target_url": audit.target_url,
            "target_provider": audit.target_provider,
            "target_model": audit.target_model,
            "status": audit.status,
            "risk_score": audit.risk_score,
            "risk_level": audit.risk_level,
            "total_tests": audit.total_tests,
            "tests_passed": audit.tests_passed,
            "vulnerabilities_found": audit.vulnerabilities_found,
            "summary": audit.summary,
            "reviewer_name": audit.reviewer_name,
            "created_at": audit.created_at.isoformat() if audit.created_at else None,
            "completed_at": audit.completed_at.isoformat() if audit.completed_at else None,
        },
        "findings": [
            {
                "id": f.id,
                "scanner_name": f.scanner_name,
                "attack_name": f.attack_name,
                "attack_prompt": f.attack_prompt,
                "target_response": f.target_response,
                "vulnerable": f.vulnerable,
                "severity": f.severity,
                "confidence": f.confidence,
                "evidence": f.evidence,
                "explanation": f.explanation,
                "recommendation": f.recommendation,
                "category": f.category,
            }
            for f in findings
        ],
        "exported_at": datetime.utcnow().isoformat(),
        "tool": "Enterprise LLM Security Auditor v1.0",
    }

    json_bytes = json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")
    filename = f"audit_{audit_id[:8]}_{audit.company_name.replace(' ', '_')}.json"
    return Response(
        content=json_bytes,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{audit_id}/export/csv")
async def export_csv(audit_id: str, session: AsyncSession = Depends(get_session)):
    audit, findings = await _get_audit_and_findings(audit_id, session)

    output = io.StringIO()
    writer = csv.writer(output)

    # Header rows
    writer.writerow(["Enterprise LLM Security Auditor — Export"])
    writer.writerow(["Company", audit.company_name])
    writer.writerow(["Audit ID", audit.id])
    writer.writerow(["Risk Score", f"{audit.risk_score:.1f}" if audit.risk_score else "N/A"])
    writer.writerow(["Risk Level", audit.risk_level or "N/A"])
    writer.writerow(["Date", audit.created_at.strftime("%Y-%m-%d") if audit.created_at else ""])
    writer.writerow([])

    # Findings
    writer.writerow([
        "Scanner", "Attack Name", "Vulnerable", "Severity", "Confidence (%)",
        "Category", "Evidence", "Explanation", "Recommendation", "Attack Prompt", "Target Response"
    ])
    for f in findings:
        writer.writerow([
            f.scanner_name, f.attack_name, "YES" if f.vulnerable else "no",
            f.severity.upper() if f.vulnerable else f.severity,
            f.confidence, f.category,
            (f.evidence or "")[:200],
            (f.explanation or "")[:300],
            (f.recommendation or "")[:300],
            (f.attack_prompt or "")[:200],
            (f.target_response or "")[:200],
        ])

    csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    filename = f"audit_{audit_id[:8]}_{audit.company_name.replace(' ', '_')}.csv"
    return Response(
        content=csv_bytes,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{audit_id}/export/xlsx")
async def export_xlsx(audit_id: str, session: AsyncSession = Depends(get_session)):
    audit, findings = await _get_audit_and_findings(audit_id, session)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl not installed")

    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E293B")

    summary_rows = [
        ("Company", audit.company_name),
        ("Audit ID", audit.id),
        ("Target URL", audit.target_url),
        ("Provider / Model", f"{audit.target_provider} / {audit.target_model}"),
        ("Risk Score", f"{audit.risk_score:.1f} / 100" if audit.risk_score else "N/A"),
        ("Risk Level", audit.risk_level or "N/A"),
        ("Total Tests", audit.total_tests),
        ("Vulnerabilities Found", audit.vulnerabilities_found),
        ("Reviewer", audit.reviewer_name),
        ("Date", audit.created_at.strftime("%Y-%m-%d %H:%M UTC") if audit.created_at else ""),
        ("Summary", audit.summary or ""),
    ]

    for i, (label, value) in enumerate(summary_rows, start=1):
        cell_label = ws_summary.cell(row=i, column=1, value=label)
        cell_label.font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 70

    # ── Findings sheet ─────────────────────────────────────────────────────────
    ws_findings = wb.create_sheet("Findings")
    headers = [
        "Scanner", "Attack Name", "Vulnerable", "Severity", "Confidence (%)",
        "Category", "Evidence", "Explanation", "Recommendation"
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws_findings.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    severity_colors = {
        "critical": "C0392B", "high": "E67E22",
        "medium": "F39C12", "low": "27AE60", "info": "3498DB"
    }

    for row_idx, f in enumerate(findings, start=2):
        row_data = [
            f.scanner_name, f.attack_name,
            "YES" if f.vulnerable else "no",
            f.severity, f.confidence, f.category,
            f.evidence or "", f.explanation or "", f.recommendation or "",
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws_findings.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        if f.vulnerable:
            color = severity_colors.get(f.severity.lower(), "333333")
            sev_cell = ws_findings.cell(row=row_idx, column=4)
            sev_cell.fill = PatternFill("solid", fgColor=color)
            sev_cell.font = Font(color="FFFFFF", bold=True)

    col_widths = [20, 30, 12, 12, 14, 20, 40, 50, 50]
    for col, width in enumerate(col_widths, start=1):
        ws_findings.column_dimensions[get_column_letter(col)].width = width

    ws_findings.freeze_panes = "A2"

    # ── Save to bytes ──────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"audit_{audit_id[:8]}_{audit.company_name.replace(' ', '_')}.xlsx"
    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
